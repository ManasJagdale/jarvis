"""
file_analysis.py

Extraction and targeted editing for document types read_file can't handle
on its own. Two categories of tool live here:

  1. Read-only extraction: analyze_file() -- pulls text/structure out of
     PDF and Excel files and hands it back RAW. This module doesn't decide
     what matters in the content -- that's the LLM's job, per-request,
     based on whatever the user actually asked for. All this does is get
     the content out of a binary format into something the LLM can read.

  2. Targeted in-place edits: edit_excel_cells() and edit_file_text().
     Both are destructive (they modify a real file on disk) and BOTH are
     unconditionally confirmation-gated in jarvis_core.py's
     ALWAYS_CONFIRM_TOOLS -- no exceptions, including for Jarvis's own
     source files.

PPTX/DOCX extraction and DOCX/XLSX generation are a deliberately deferred
follow-up phase -- not built here. See the project spec for the reasoning.

------------------------------------------------------------------------
ONE-TIME SETUP -- Tesseract OCR (needed only for scanned/no-text-layer PDFs):

  Most PDFs (anything with a real, selectable text layer) need NO extra
  setup at all -- PyMuPDF extracts that text directly. OCR is only a
  fallback for PDFs that are actually scanned images with no text layer.

  1. Download the Windows installer from:
     https://github.com/UB-Mannheim/tesseract/wiki
  2. Run it (default install path is fine:
     C:\\Program Files\\Tesseract-OCR\\tesseract.exe)
  3. Add that folder to your PATH, OR set it explicitly in code with:
         pytesseract.pytesseract.tesseract_cmd = r"C:\\Program Files\\Tesseract-OCR\\tesseract.exe"
     (add that line near the top of this file if PATH doesn't pick it up)

  If Tesseract isn't installed, analyze_file() on a scanned PDF returns a
  clear error saying so, rather than crashing -- text-layer PDFs work fine
  either way.

Install the Python side with:
    pip install PyMuPDF pytesseract openpyxl pillow
------------------------------------------------------------------------
"""

import os
import re

import fitz  # PyMuPDF
import openpyxl
import pytesseract
from PIL import Image

from config import MAX_ANALYZE_CHARS

# Crude, deliberately best-effort signal for "this OCR word looks like it's
# part of a formula" -- catches common math operators/symbols. This is NOT
# real formula-region detection (that needs an actual layout model); it
# will miss subtler equations and occasionally over-capture plain text
# that happens to contain one of these characters. Documented as such in
# every place this gets surfaced to the user.
_MATH_SYMBOL_PATTERN = re.compile(r"[=+\-*/^√∫∑π≈≤≥]")


def _tesseract_available() -> bool:
    try:
        pytesseract.get_tesseract_version()
        return True
    except Exception:
        return False


def _truncate(text: str, extra_note: str = "") -> str:
    if len(text) <= MAX_ANALYZE_CHARS:
        return text
    remaining = len(text) - MAX_ANALYZE_CHARS
    return (
        text[:MAX_ANALYZE_CHARS]
        + f"\n\n[...truncated, {remaining} more characters not shown{extra_note}]"
    )


def analyze_file(path: str, extract_visuals: bool = False, output_dir: str = "") -> str:
    """
    Extract raw text/structure from a PDF or Excel file. extract_visuals
    (PDF only) additionally crops and saves diagrams and formula-looking
    regions as images to output_dir -- off by default since scanning every
    page for visual regions costs real time you don't want to pay on a
    quick text-only read.
    """
    if not os.path.exists(path):
        return f"Error: file not found at {path}"

    ext = os.path.splitext(path)[1].lower()

    try:
        if ext == ".pdf":
            return _analyze_pdf(path, extract_visuals, output_dir)
        if ext in (".xlsx", ".xlsm"):
            return _analyze_excel(path)
        return (
            f"Error: analyze_file doesn't support '{ext}' files yet -- "
            "currently PDF and Excel (.xlsx) only."
        )
    except Exception as exc:  # noqa: BLE001 - never let a bad file crash the turn
        return f"Error analyzing file: {exc}"


def _analyze_pdf(path: str, extract_visuals: bool, output_dir: str) -> str:
    if extract_visuals and not output_dir:
        return "Error: extract_visuals=True requires output_dir -- where should diagrams/formulas be saved?"

    try:
        doc = fitz.open(path)
    except Exception as exc:
        return f"Error opening PDF: {exc}"

    if doc.is_encrypted:
        doc.close()
        return "Error: this PDF appears password-protected -- Jarvis can't open protected PDFs."

    if extract_visuals:
        os.makedirs(output_dir, exist_ok=True)

    text_parts = []
    ocr_used = False
    visual_count = 0

    for page_num, page in enumerate(doc, start=1):
        page_text = page.get_text().strip()

        if not page_text:
            if not _tesseract_available():
                doc.close()
                return (
                    "Error: this PDF has no extractable text layer (looks scanned), "
                    "and Tesseract OCR isn't installed -- see the setup steps in "
                    "file_analysis.py's docstring."
                )
            ocr_used = True
            pix = page.get_pixmap(dpi=200)
            img = Image.frombytes("RGB", (pix.width, pix.height), pix.samples)
            page_text = pytesseract.image_to_string(img).strip()

        text_parts.append(f"--- Page {page_num} ---\n{page_text}")

        if extract_visuals:
            visual_count += _extract_page_diagrams(doc, page, page_num, output_dir)
            if _tesseract_available():
                visual_count += _extract_page_formula_regions(page, page_num, output_dir)

    doc.close()

    result = _truncate("\n\n".join(text_parts))

    prefix = ""
    if ocr_used:
        prefix += (
            "(At least one page had no text layer, so OCR was used for it -- "
            "expect occasional errors in that page's text.)\n\n"
        )
    if extract_visuals:
        prefix += (
            f"(Saved {visual_count} diagram/formula image(s) to {output_dir}. "
            "Formula-region detection is a rough heuristic, not real formula "
            "recognition -- it may miss some equations or over-capture plain "
            "text containing math symbols.)\n\n"
        )

    return prefix + result


def _extract_page_diagrams(doc, page, page_num: int, output_dir: str) -> int:
    """Save embedded raster images on this page (diagrams/figures) as files."""
    saved = 0
    for img_index, img_info in enumerate(page.get_images(full=True), start=1):
        xref = img_info[0]
        try:
            base_image = doc.extract_image(xref)
            out_path = os.path.join(
                output_dir, f"page{page_num}_diagram{img_index}.{base_image['ext']}"
            )
            with open(out_path, "wb") as f:
                f.write(base_image["image"])
            saved += 1
        except Exception:
            continue  # a single bad embedded image shouldn't abort the whole extraction
    return saved


def _extract_page_formula_regions(page, page_num: int, output_dir: str) -> int:
    """
    Best-effort: render the page, OCR it for word bounding boxes, group
    words into their OCR-detected lines, and crop any LINE that contains
    at least one math-symbol word -- one crop per matching line, not one
    per matching word (which would produce several overlapping crops for
    a single equation). See _MATH_SYMBOL_PATTERN's comment above for the
    honest limitations of this approach.
    """
    saved = 0
    pix = page.get_pixmap(dpi=200)
    img = Image.frombytes("RGB", (pix.width, pix.height), pix.samples)
    data = pytesseract.image_to_data(img, output_type=pytesseract.Output.DICT)

    lines: dict[tuple, list[int]] = {}
    for i, word in enumerate(data["text"]):
        if not word.strip():
            continue
        key = (data["block_num"][i], data["par_num"][i], data["line_num"][i])
        lines.setdefault(key, []).append(i)

    for (block_num, par_num, line_num), idxs in lines.items():
        if not any(_MATH_SYMBOL_PATTERN.search(data["text"][i]) for i in idxs):
            continue

        xs = [data["left"][i] for i in idxs]
        ys = [data["top"][i] for i in idxs]
        rights = [data["left"][i] + data["width"][i] for i in idxs]
        bottoms = [data["top"][i] + data["height"][i] for i in idxs]
        pad_x, pad_y = 30, 15
        box = (
            max(0, min(xs) - pad_x),
            max(0, min(ys) - pad_y),
            min(img.width, max(rights) + pad_x),
            min(img.height, max(bottoms) + pad_y),
        )
        try:
            crop = img.crop(box)
            out_path = os.path.join(
                output_dir, f"page{page_num}_formula_line{line_num}_{par_num}.png"
            )
            crop.save(out_path)
            saved += 1
        except Exception:
            continue
    return saved


def _analyze_excel(path: str) -> str:
    try:
        wb_values = openpyxl.load_workbook(path, data_only=True)
        wb_formulas = openpyxl.load_workbook(path, data_only=False)
    except Exception as exc:
        return f"Error opening Excel file: {exc}"

    sections = []
    for sheet_name in wb_values.sheetnames:
        ws_values = wb_values[sheet_name]
        ws_formulas = wb_formulas[sheet_name]

        header = f"=== Sheet: {sheet_name} ({ws_values.max_row} rows x {ws_values.max_column} cols) ==="

        formula_lines = []
        for row in ws_formulas.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    computed = ws_values[cell.coordinate].value
                    if computed is None:
                        computed_note = "(not cached -- open/save this file in Excel once to compute it)"
                    else:
                        computed_note = computed
                    formula_lines.append(f"{cell.coordinate}: {cell.value} -> {computed_note}")
        formula_block = "Formulas:\n" + "\n".join(formula_lines) if formula_lines else "Formulas: (none)"

        row_lines = [
            ", ".join("" if v is None else str(v) for v in row)
            for row in ws_values.iter_rows(values_only=True)
        ]
        data_block = "Data:\n" + "\n".join(row_lines)

        sections.append(f"{header}\n{formula_block}\n\n{data_block}")

    full_text = "\n\n".join(sections)
    return _truncate(full_text, extra_note=" -- ask for a specific sheet or row range instead")


def edit_excel_cells(path: str, edits: dict) -> str:
    """
    Apply in-place cell edits to an existing Excel file. `edits` maps a
    cell coordinate to its new value/formula, e.g. {"C4": "=B4*1.05"}.
    Prefix a coordinate with "SheetName!" to target a sheet other than
    the first (e.g. "Sheet2!C4"). Overwrites the file directly -- this is
    why it's unconditionally confirmation-gated in jarvis_core.py.
    """
    if not os.path.exists(path):
        return f"Error: file not found at {path}"

    try:
        wb = openpyxl.load_workbook(path, data_only=False)
    except Exception as exc:
        return f"Error opening Excel file: {exc}"

    applied = []
    for coord, new_value in edits.items():
        if "!" in coord:
            sheet_name, cell_ref = coord.split("!", 1)
            if sheet_name not in wb.sheetnames:
                return f"Error: sheet '{sheet_name}' not found in {path} -- no changes made."
            ws = wb[sheet_name]
        else:
            cell_ref = coord
            ws = wb[wb.sheetnames[0]]
        ws[cell_ref] = new_value
        applied.append(f"{ws.title}!{cell_ref} = {new_value}")

    try:
        wb.save(path)
    except Exception as exc:
        return f"Error saving Excel file (changes NOT applied): {exc}"

    return "Updated:\n" + "\n".join(applied)


def edit_file_text(path: str, old_text: str, new_text: str) -> str:
    """
    Exact find-and-replace on a text file (e.g. a .py file). old_text must
    appear EXACTLY ONCE in the file -- fails loudly rather than guessing
    which occurrence was meant if it's not unique, and fails loudly if
    it's not found at all rather than silently doing nothing.
    """
    if not os.path.exists(path):
        return f"Error: file not found at {path}"

    try:
        with open(path, "r", encoding="utf-8") as f:
            content = f.read()
    except Exception as exc:
        return f"Error reading file: {exc}"

    count = content.count(old_text)
    if count == 0:
        return f"Error: old_text not found in {path} -- no changes made."
    if count > 1:
        return f"Error: old_text matches {count} places in {path} -- must be unique. No changes made."

    new_content = content.replace(old_text, new_text, 1)
    try:
        with open(path, "w", encoding="utf-8") as f:
            f.write(new_content)
    except Exception as exc:
        return f"Error writing file: {exc}"

    return f"Edited {path} -- replaced 1 occurrence."
